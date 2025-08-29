#!/usr/bin/env python3
"""
OVN-Kubernetes Benchmark MCP Agent - Report Generation
AI agent using LangGraph to generate performance reports with historical analysis
"""

import asyncio
import json
import os
import uuid
from datetime import datetime, timezone, timedelta
from typing import Dict, List, Any, Optional, TypedDict
import pandas as pd
import aiohttp
from pathlib import Path

from langgraph.graph import StateGraph, END
from langgraph.checkpoint.memory import MemorySaver
from langchain.schema import HumanMessage, AIMessage, SystemMessage
from langchain_openai import ChatOpenAI

# Report generation imports
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.charts import BarChart, LineChart, Reference
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import matplotlib.pyplot as plt
import seaborn as sns


class ReportAgentState(TypedDict):
    """State for the report generation agent"""
    messages: List[Any]
    current_step: str
    historical_data: Dict[str, Any]
    report_data: Dict[str, Any]
    analysis: Dict[str, Any]
    errors: List[str]
    report_period: int
    run_id: str
    timestamp: str
    config: Dict[str, Any]
    output_files: List[str]


class PerformanceReportAgent:
    """AI agent for generating performance reports with historical analysis"""
    
    def __init__(self, 
                 mcp_server_url: str = "http://localhost:8000",
                 openai_api_key: Optional[str] = None,
                 report_period_days: int = 7,
                 output_dir: str = "exports"):
        self.mcp_server_url = mcp_server_url.rstrip('/')
        self.report_period_days = report_period_days
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.run_id = str(uuid.uuid4())
        
        # Initialize LLM
        self.llm = ChatOpenAI(
            model="gpt-4o-mini",
            temperature=0.3,
            api_key=openai_api_key
        )
        
        # Create workflow graph
        self.workflow = self._create_workflow()
        self.app = self.workflow.compile(checkpointer=MemorySaver())
    
    def _create_workflow(self) -> StateGraph:
        """Create the LangGraph workflow for report generation"""
        workflow = StateGraph(ReportAgentState)
        
        # Add nodes
        workflow.add_node("initialize", self._initialize_report)
        workflow.add_node("fetch_historical_data", self._fetch_historical_data)
        workflow.add_node("analyze_performance", self._analyze_performance)
        workflow.add_node("generate_insights", self._generate_insights)
        workflow.add_node("create_excel_report", self._create_excel_report)
        workflow.add_node("create_pdf_report", self._create_pdf_report)
        workflow.add_node("finalize_report", self._finalize_report)
        workflow.add_node("handle_error", self._handle_error)
        
        # Set entry point
        workflow.set_entry_point("initialize")
        
        # Add edges
        workflow.add_edge("initialize", "fetch_historical_data")
        workflow.add_edge("fetch_historical_data", "analyze_performance")
        workflow.add_edge("analyze_performance", "generate_insights")
        workflow.add_edge("generate_insights", "create_excel_report")
        workflow.add_edge("create_excel_report", "create_pdf_report")
        workflow.add_edge("create_pdf_report", "finalize_report")
        workflow.add_edge("handle_error", "finalize_report")
        workflow.add_edge("finalize_report", END)
        
        return workflow
    
    async def _initialize_report(self, state: ReportAgentState) -> ReportAgentState:
        """Initialize the report generation process"""
        print(f"📊 Starting performance report generation - Run ID: {self.run_id}")
        
        state["current_step"] = "initialize"
        state["run_id"] = self.run_id
        state["timestamp"] = datetime.now(timezone.utc).isoformat()
        state["report_period"] = self.report_period_days
        state["historical_data"] = {}
        state["report_data"] = {}
        state["analysis"] = {}
        state["errors"] = []
        state["output_files"] = []
        state["config"] = {
            "mcp_server_url": self.mcp_server_url,
            "report_period_days": self.report_period_days,
            "output_directory": str(self.output_dir)
        }
        
        # Add initialization message
        state["messages"] = [
            SystemMessage(content="""You are an AI performance report analyst for OpenShift OVN-Kubernetes clusters.
            Your task is to analyze historical performance data, identify trends, generate insights, and create comprehensive reports.
            
            Your analysis should include:
            1. Performance trends over time
            2. Health score analysis
            3. Alert patterns and frequencies
            4. Resource utilization trends
            5. Recommendations for optimization
            6. Comparison with previous periods
            
            Generate actionable insights and recommendations based on the data."""),
            HumanMessage(content=f"Generate a comprehensive performance report for the last {self.report_period_days} days")
        ]
        
        return state
    
    async def _fetch_historical_data(self, state: ReportAgentState) -> ReportAgentState:
        """Fetch historical performance data"""
        print(f"📈 Fetching historical data for the last {self.report_period_days} days...")
        state["current_step"] = "fetch_historical_data"
        
        try:
            # Fetch overall performance history
            overall_data = await self._call_mcp_tool("get_performance_history", {
                "days": self.report_period_days
            })
            state["historical_data"]["overall"] = json.loads(overall_data)
            
            # Fetch category-specific data
            categories = ["api_server", "multus", "ovnk_pods", "ovnk_containers", "ovnk_sync"]
            for category in categories:
                try:
                    category_data = await self._call_mcp_tool("get_performance_history", {
                        "days": self.report_period_days,
                        "metric_type": category
                    })
                    state["historical_data"][category] = json.loads(category_data)
                except Exception as e:
                    state["errors"].append(f"Failed to fetch {category} data: {str(e)}")
            
            state["messages"].append(
                AIMessage(content=f"✅ Successfully fetched historical data for {len(state['historical_data'])} categories")
            )
            
        except Exception as e:
            error_msg = f"Failed to fetch historical data: {str(e)}"
            state["errors"].append(error_msg)
            state["messages"].append(AIMessage(content=f"❌ {error_msg}"))
        
        return state
    
    async def _analyze_performance(self, state: ReportAgentState) -> ReportAgentState:
        """Analyze performance data and calculate trends"""
        print("🔍 Analyzing performance trends...")
        state["current_step"] = "analyze_performance"
        
        try:
            analysis = {}
            
            # Analyze overall trends
            if "overall" in state["historical_data"]:
                overall_data = state["historical_data"]["overall"]["data"]
                
                # Overall trend analysis
                if "overall_trend" in overall_data:
                    trend_data = overall_data["overall_trend"]
                    if trend_data:
                        health_scores = [item["avg_health"] for item in trend_data if item["avg_health"] > 0]
                        total_alerts = [item["total_alerts"] for item in trend_data]
                        
                        analysis["overall_trends"] = {
                            "avg_health_score": sum(health_scores) / len(health_scores) if health_scores else 0,
                            "health_trend": self._calculate_trend(health_scores),
                            "total_alerts": sum(total_alerts),
                            "alert_trend": self._calculate_trend(total_alerts),
                            "days_analyzed": len(trend_data)
                        }
                
                # Cluster growth analysis
                if "cluster_trend" in overall_data:
                    cluster_data = overall_data["cluster_trend"]
                    if cluster_data:
                        analysis["cluster_growth"] = {
                            "avg_networkpolicies": sum(item["avg_networkpolicies"] for item in cluster_data) / len(cluster_data),
                            "avg_nodes": sum(item["max_nodes"] for item in cluster_data) / len(cluster_data),
                            "avg_namespaces": sum(item["max_namespaces"] for item in cluster_data) / len(cluster_data)
                        }
            
            # Analyze category-specific performance
            categories = ["api_server", "multus", "ovnk_pods", "ovnk_containers", "ovnk_sync"]
            analysis["category_analysis"] = {}
            
            for category in categories:
                if category in state["historical_data"] and f"{category}_specific" in state["historical_data"][category]["data"]:
                    category_data = state["historical_data"][category]["data"][f"{category}_specific"]
                    
                    if category_data:
                        health_scores = [item["health_score"] for item in category_data if item["health_score"]]
                        alert_counts = [len(item["alerts"]) for item in category_data]
                        
                        analysis["category_analysis"][category] = {
                            "avg_health_score": sum(health_scores) / len(health_scores) if health_scores else 0,
                            "health_trend": self._calculate_trend(health_scores),
                            "total_alerts": sum(alert_counts),
                            "alert_frequency": sum(alert_counts) / len(alert_counts) if alert_counts else 0,
                            "measurements": len(category_data)
                        }
            
            state["analysis"] = analysis
            state["messages"].append(
                AIMessage(content="✅ Performance analysis completed with trend calculations")
            )
            
        except Exception as e:
            error_msg = f"Failed to analyze performance data: {str(e)}"
            state["errors"].append(error_msg)
            state["messages"].append(AIMessage(content=f"❌ {error_msg}"))
        
        return state
    
    async def _generate_insights(self, state: ReportAgentState) -> ReportAgentState:
        """Generate AI-powered insights and recommendations"""
        print("🧠 Generating AI insights and recommendations...")
        state["current_step"] = "generate_insights"
        
        try:
            # Prepare data for LLM analysis
            analysis_context = {
                "report_period": state["report_period"],
                "analysis_results": state["analysis"],
                "error_count": len(state["errors"])
            }
            
            # Generate insights using LLM
            insight_prompt = f"""
            Based on the following OpenShift OVN-Kubernetes performance analysis for the last {state['report_period']} days:
            
            {json.dumps(analysis_context, indent=2)}
            
            Please provide:
            1. Key performance insights and trends
            2. Areas of concern that need attention
            3. Performance improvements observed
            4. Specific recommendations for optimization
            5. Risk assessment for cluster health
            
            Format your response as JSON with the following structure:
            {{
                "key_insights": ["insight1", "insight2", ...],
                "concerns": ["concern1", "concern2", ...],
                "improvements": ["improvement1", "improvement2", ...],
                "recommendations": [
                    {{"priority": "high|medium|low", "category": "category", "description": "desc", "impact": "impact"}},
                    ...
                ],
                "risk_assessment": {{"level": "low|medium|high", "description": "desc"}},
                "executive_summary": "summary"
            }}
            """
            
            response = await self.llm.ainvoke([
                SystemMessage(content="You are an expert OpenShift performance analyst."),
                HumanMessage(content=insight_prompt)
            ])
            
            # Parse AI response
            try:
                insights = json.loads(response.content)
                state["analysis"]["ai_insights"] = insights
            except json.JSONDecodeError:
                # Fallback to basic insights if JSON parsing fails
                state["analysis"]["ai_insights"] = {
                    "key_insights": ["Analysis completed successfully"],
                    "concerns": [],
                    "improvements": [],
                    "recommendations": [],
                    "risk_assessment": {"level": "medium", "description": "Unable to parse AI insights"},
                    "executive_summary": response.content[:500]
                }
            
            state["messages"].append(
                AIMessage(content="✅ AI-powered insights and recommendations generated")
            )
            
        except Exception as e:
            error_msg = f"Failed to generate insights: {str(e)}"
            state["errors"].append(error_msg)
            state["messages"].append(AIMessage(content=f"❌ {error_msg}"))
        
        return state
    
    async def _create_excel_report(self, state: ReportAgentState) -> ReportAgentState:
        """Create Excel report with charts and tables"""
        print("📄 Creating Excel report...")
        state["current_step"] = "create_excel_report"
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ovnk_performance_report_{timestamp}.xlsx"
            filepath = self.output_dir / filename
            
            wb = Workbook()
            
            # Remove default sheet
            wb.remove(wb.active)
            
            # Executive Summary Sheet
            self._create_excel_summary_sheet(wb, state)
            
            # Historical Trends Sheet
            self._create_excel_trends_sheet(wb, state)
            
            # Category Analysis Sheet
            self._create_excel_category_sheet(wb, state)
            
            # Recommendations Sheet
            self._create_excel_recommendations_sheet(wb, state)
            
            # Raw Data Sheet
            self._create_excel_raw_data_sheet(wb, state)
            
            wb.save(filepath)
            state["output_files"].append(str(filepath))
            
            state["messages"].append(
                AIMessage(content=f"✅ Excel report created: {filename}")
            )
            
        except Exception as e:
            error_msg = f"Failed to create Excel report: {str(e)}"
            state["errors"].append(error_msg)
            state["messages"].append(AIMessage(content=f"❌ {error_msg}"))
        
        return state
    
    async def _create_pdf_report(self, state: ReportAgentState) -> ReportAgentState:
        """Create PDF summary report"""
        print("📋 Creating PDF summary report...")
        state["current_step"] = "create_pdf_report"
        
        try:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"ovnk_performance_summary_{timestamp}.pdf"
            filepath = self.output_dir / filename
            
            # Create PDF
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # Title
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Title'],
                fontSize=20,
                spaceAfter=30,
                textColor=colors.darkblue
            )
            
            story.append(Paragraph("OpenShift OVN-Kubernetes Performance Report", title_style))
            story.append(Spacer(1, 20))
            
            # Executive Summary
            story.append(Paragraph("Executive Summary", styles['Heading1']))
            
            if "ai_insights" in state["analysis"]:
                insights = state["analysis"]["ai_insights"]
                summary = insights.get("executive_summary", "No executive summary available")
                story.append(Paragraph(summary, styles['Normal']))
            
            story.append(Spacer(1, 15))
            
            # Key Metrics Table
            story.append(Paragraph("Key Performance Metrics", styles['Heading2']))
            
            if "overall_trends" in state["analysis"]:
                trends = state["analysis"]["overall_trends"]
                metrics_data = [
                    ["Metric", "Value"],
                    ["Average Health Score", f"{trends.get('avg_health_score', 0):.1f}"],
                    ["Health Trend", trends.get('health_trend', 'stable').title()],
                    ["Total Alerts", str(trends.get('total_alerts', 0))],
                    ["Days Analyzed", str(trends.get('days_analyzed', 0))]
                ]
                
                metrics_table = Table(metrics_data)
                metrics_table.setStyle(TableStyle([
                    ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 14),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('GRID', (0, 0), (-1, -1), 1, colors.black)
                ]))
                
                story.append(metrics_table)
            
            story.append(PageBreak())
            
            # Category Analysis
            story.append(Paragraph("Category Performance Analysis", styles['Heading1']))
            
            if "category_analysis" in state["analysis"]:
                for category, analysis in state["analysis"]["category_analysis"].items():
                    story.append(Paragraph(f"{category.replace('_', ' ').title()}", styles['Heading2']))
                    
                    category_data = [
                        ["Metric", "Value"],
                        ["Average Health Score", f"{analysis.get('avg_health_score', 0):.1f}"],
                        ["Health Trend", analysis.get('health_trend', 'stable').title()],
                        ["Total Alerts", str(analysis.get('total_alerts', 0))],
                        ["Measurements", str(analysis.get('measurements', 0))]
                    ]
                    
                    category_table = Table(category_data)
                    category_table.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.black),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('GRID', (0, 0), (-1, -1), 1, colors.black)
                    ]))
                    
                    story.append(category_table)
                    story.append(Spacer(1, 10))
            
            story.append(PageBreak())
            
            # Recommendations
            if "ai_insights" in state["analysis"] and "recommendations" in state["analysis"]["ai_insights"]:
                story.append(Paragraph("Recommendations", styles['Heading1']))
                
                recommendations = state["analysis"]["ai_insights"]["recommendations"]
                for i, rec in enumerate(recommendations, 1):
                    story.append(Paragraph(f"{i}. {rec.get('description', 'No description')}", styles['Normal']))
                    story.append(Paragraph(f"Priority: {rec.get('priority', 'medium').title()}", styles['Italic']))
                    story.append(Spacer(1, 10))
            
            # Build PDF
            doc.build(story)
            state["output_files"].append(str(filepath))
            
            state["messages"].append(
                AIMessage(content=f"✅ PDF summary report created: {filename}")
            )
            
        except Exception as e:
            error_msg = f"Failed to create PDF report: {str(e)}"
            state["errors"].append(error_msg)
            state["messages"].append(AIMessage(content=f"❌ {error_msg}"))
        
        return state
    
    async def _finalize_report(self, state: ReportAgentState) -> ReportAgentState:
        """Finalize the report generation process"""
        print("🎯 Finalizing report generation...")
        state["current_step"] = "finalize"
        
        # Print report summary to stdout
        print("\n" + "="*80)
        print("PERFORMANCE REPORT SUMMARY")
        print("="*80)
        
        if "overall_trends" in state["analysis"]:
            trends = state["analysis"]["overall_trends"]
            print(f"📊 Overall Health Score: {trends.get('avg_health_score', 0):.1f}/100")
            print(f"📈 Health Trend: {trends.get('health_trend', 'stable').title()}")
            print(f"🚨 Total Alerts: {trends.get('total_alerts', 0)}")
            print(f"📅 Days Analyzed: {trends.get('days_analyzed', 0)}")
        
        print(f"\n📁 Generated Files:")
        for file_path in state["output_files"]:
            print(f"   • {file_path}")
        
        if state["errors"]:
            print(f"\n⚠️ Errors encountered: {len(state['errors'])}")
            for error in state["errors"]:
                print(f"   • {error}")
        
        print("="*80)
        
        # Generate final summary message
        final_msg = f"""
🎯 Performance Report Generation Complete - Run ID: {state['run_id']}

📊 Report Summary:
• Report Period: {state['report_period']} days
• Files Generated: {len(state['output_files'])}
• Total Errors: {len(state['errors'])}
• Timestamp: {state['timestamp']}

📁 Output Files:
{chr(10).join(f'• {os.path.basename(f)}' for f in state['output_files'])}
"""
        
        if "ai_insights" in state["analysis"]:
            insights = state["analysis"]["ai_insights"]
            final_msg += f"""
🧠 AI Insights Summary:
• Key Insights: {len(insights.get('key_insights', []))}
• Concerns Identified: {len(insights.get('concerns', []))}
• Recommendations: {len(insights.get('recommendations', []))}
• Risk Level: {insights.get('risk_assessment', {}).get('level', 'unknown').title()}
"""
        
        state["messages"].append(AIMessage(content=final_msg))
        
        print("✅ Performance report generation completed successfully!")
        
        return state
    
    async def _handle_error(self, state: ReportAgentState) -> ReportAgentState:
        """Handle errors during report generation"""
        state["current_step"] = "handle_error"
        error_msg = f"Critical error in step {state['current_step']}"
        state["errors"].append(error_msg)
        state["messages"].append(AIMessage(content=f"🚨 {error_msg}"))
        return state
    
    def _calculate_trend(self, values: List[float]) -> str:
        """Calculate trend direction from a list of values"""
        if len(values) < 2:
            return "insufficient_data"
        
        # Simple linear trend calculation
        first_half = values[:len(values)//2]
        second_half = values[len(values)//2:]
        
        first_avg = sum(first_half) / len(first_half)
        second_avg = sum(second_half) / len(second_half)
        
        diff_pct = ((second_avg - first_avg) / first_avg * 100) if first_avg != 0 else 0
        
        if diff_pct > 5:
            return "improving"
        elif diff_pct < -5:
            return "declining"
        else:
            return "stable"
    
    def _create_excel_summary_sheet(self, wb: Workbook, state: ReportAgentState) -> None:
        """Create executive summary sheet in Excel"""
        ws = wb.create_sheet("Executive Summary", 0)
        
        # Headers
        ws['A1'] = "OpenShift OVN-Kubernetes Performance Report"
        ws['A1'].font = Font(size=16, bold=True)
        
        ws['A3'] = f"Report Period: {state['report_period']} days"
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S UTC')}"
        
        # Key metrics
        row = 6
        if "overall_trends" in state["analysis"]:
            trends = state["analysis"]["overall_trends"]
            ws[f'A{row}'] = "Key Performance Metrics"
            ws[f'A{row}'].font = Font(bold=True)
            row += 2
            
            metrics = [
                ("Average Health Score", f"{trends.get('avg_health_score', 0):.1f}"),
                ("Health Trend", trends.get('health_trend', 'stable').title()),
                ("Total Alerts", str(trends.get('total_alerts', 0))),
                ("Days Analyzed", str(trends.get('days_analyzed', 0)))
            ]
            
            for metric, value in metrics:
                ws[f'A{row}'] = metric
                ws[f'B{row}'] = value
                row += 1
    
    def _create_excel_trends_sheet(self, wb: Workbook, state: ReportAgentState) -> None:
        """Create trends analysis sheet in Excel"""
        ws = wb.create_sheet("Historical Trends")
        
        ws['A1'] = "Historical Performance Trends"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add trend data if available
        row = 3
        if "overall" in state["historical_data"] and "overall_trend" in state["historical_data"]["overall"]["data"]:
            trend_data = state["historical_data"]["overall"]["data"]["overall_trend"]
            
            if trend_data:
                headers = ["Date", "Avg Health", "Min Health", "Max Health", "Total Alerts"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=header).font = Font(bold=True)
                
                row += 1
                for item in trend_data:
                    ws.cell(row=row, column=1, value=str(item.get("date", "")))
                    ws.cell(row=row, column=2, value=item.get("avg_health", 0))
                    ws.cell(row=row, column=3, value=item.get("min_health", 0))
                    ws.cell(row=row, column=4, value=item.get("max_health", 0))
                    ws.cell(row=row, column=5, value=item.get("total_alerts", 0))
                    row += 1
    
    def _create_excel_category_sheet(self, wb: Workbook, state: ReportAgentState) -> None:
        """Create category analysis sheet in Excel"""
        ws = wb.create_sheet("Category Analysis")
        
        ws['A1'] = "Category Performance Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        if "category_analysis" in state["analysis"]:
            headers = ["Category", "Avg Health", "Health Trend", "Total Alerts", "Measurements"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            
            row += 1
            for category, analysis in state["analysis"]["category_analysis"].items():
                ws.cell(row=row, column=1, value=category.replace('_', ' ').title())
                ws.cell(row=row, column=2, value=analysis.get('avg_health_score', 0))
                ws.cell(row=row, column=3, value=analysis.get('health_trend', 'stable').title())
                ws.cell(row=row, column=4, value=analysis.get('total_alerts', 0))
                ws.cell(row=row, column=5, value=analysis.get('measurements', 0))
                row += 1
    
    def _create_excel_recommendations_sheet(self, wb: Workbook, state: ReportAgentState) -> None:
        """Create recommendations sheet in Excel"""
        ws = wb.create_sheet("Recommendations")
        
        ws['A1'] = "AI-Generated Recommendations"
        ws['A1'].font = Font(size=14, bold=True)
        
        row = 3
        if "ai_insights" in state["analysis"] and "recommendations" in state["analysis"]["ai_insights"]:
            headers = ["Priority", "Category", "Description", "Impact"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=row, column=col, value=header).font = Font(bold=True)
            
            row += 1
            recommendations = state["analysis"]["ai_insights"]["recommendations"]
            for rec in recommendations:
                ws.cell(row=row, column=1, value=rec.get('priority', 'medium').title())
                ws.cell(row=row, column=2, value=rec.get('category', ''))
                ws.cell(row=row, column=3, value=rec.get('description', ''))
                ws.cell(row=row, column=4, value=rec.get('impact', ''))
                row += 1
    
    def _create_excel_raw_data_sheet(self, wb: Workbook, state: ReportAgentState) -> None:
        """Create raw data sheet in Excel"""
        ws = wb.create_sheet("Raw Data")
        
        ws['A1'] = "Raw Historical Data"
        ws['A1'].font = Font(size=14, bold=True)
        
        # Add raw data as JSON
        ws['A3'] = "Historical Data (JSON):"
        ws['A3'].font = Font(bold=True)
        
        raw_data = json.dumps(state["historical_data"], indent=2)
        ws['A5'] = raw_data
    
    async def _call_mcp_tool(self, tool_name: str, params: Dict[str, Any]) -> str:
        """Call MCP server tool"""
        async with aiohttp.ClientSession() as session:
            payload = {
                "tool": tool_name,
                "params": params
            }
            
            async with session.post(
                f"{self.mcp_server_url}/call_tool",
                json=payload,
                timeout=aiohttp.ClientTimeout(total=120)
            ) as response:
                if response.status != 200:
                    raise Exception(f"MCP server returned {response.status}: {await response.text()}")
                
                result = await response.json()
                if "error" in result:
                    raise Exception(result["error"])
                
                return result.get("content", "{}")
    
    async def generate_performance_report(self) -> Dict[str, Any]:
        """Main method to generate performance report"""
        try:
            print(f"📊 Starting AI-driven performance report generation...")
            
            # Initialize state
            initial_state = ReportAgentState(
                messages=[],
                current_step="",
                historical_data={},
                report_data={},
                analysis={},
                errors=[],
                report_period=self.report_period_days,
                run_id=self.run_id,
                timestamp=datetime.now(timezone.utc).isoformat(),
                config={},
                output_files=[]
            )
            
            # Run the workflow
            config = {"configurable": {"thread_id": self.run_id}}
            final_state = await self.app.ainvoke(initial_state, config)
            
            return {
                "run_id": final_state["run_id"],
                "timestamp": final_state["timestamp"],
                "report_period": final_state["report_period"],
                "output_files": final_state["output_files"],
                "analysis": final_state["analysis"],
                "errors": final_state["errors"],
                "success": len(final_state["errors"]) == 0,
                "summary": {
                    "files_generated": len(final_state["output_files"]),
                    "categories_analyzed": len(final_state.get("analysis", {}).get("category_analysis", {})),
                    "total_errors": len(final_state["errors"])
                }
            }
            
        except Exception as e:
            print(f"❌ Critical error in performance report generation: {e}")
            return {
                "run_id": self.run_id,
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "error": str(e),
                "success": False
            }


async def main():
    """Main function for standalone execution"""
    import os
    
    # Configuration
    mcp_server_url = os.getenv("MCP_SERVER_URL", "http://localhost:8000")
    openai_api_key = os.getenv("OPENAI_API_KEY")
    report_period = int(os.getenv("REPORT_PERIOD_DAYS", "7"))
    output_dir = os.getenv("REPORT_OUTPUT_DIR", "exports")
    
    if not openai_api_key:
        print("❌ OPENAI_API_KEY environment variable is required")
        return
    
    # Create and run agent
    agent = PerformanceReportAgent(
        mcp_server_url=mcp_server_url,
        openai_api_key=openai_api_key,
        report_period_days=report_period,
        output_dir=output_dir
    )
    
    # Generate report
    result = await agent.generate_performance_report()
    
    # Print results
    print("\n" + "="*60)
    print("PERFORMANCE REPORT GENERATION RESULTS")
    print("="*60)
    print(json.dumps(result, indent=2, default=str))


if __name__ == "__main__":
    asyncio.run(main())